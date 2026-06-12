# -*- coding: utf-8 -*-
"""
Normalización y consolidación del tablero Sedes Educativas - UESVALLE.

Entradas esperadas:
  data/sedes_educativas/raw/Escuelas_IRCAS.xlsx
  data/sedes_educativas/raw/SedesUESVALLE.xlsx
  data/sedes_educativas/raw/MEN_SEDES_EDUCATIVAS_PREESCOLAR_BASICA_Y_MEDIA_20260605.csv

Salidas:
  data/sedes_educativas/current/sedes_educativas_irca_consolidado.csv
  data/sedes_educativas/current/metadata_sedes_educativas.json
  data/sedes_educativas/current/resumen_sedes_municipio.csv
  data/sedes_educativas/current/resumen_sedes_riesgo.csv
  data/sedes_educativas/current/resumen_acueducto.csv
  data/sedes_educativas/current/control_cruce_sedes_irca.csv
  data/sedes_educativas/current/sedes_sin_georreferenciar.csv
  data/sedes_educativas/current/maestro_acueductos_mapa.csv
  data/sedes_educativas/current/validacion_matricula_fuentes.csv
  data/sedes_educativas/current/validacion_men_sedes.csv
  docs/sedes_educativas/SEDES_EDUCATIVAS_IRCA_BASE_CONSOLIDADA_CONTROL.xlsx

Versión: V3
Fecha: 2026-06-05
"""

from __future__ import annotations

import json
import math
import re
import shutil
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    print("[ERROR] Faltan librerías requeridas: pandas y openpyxl.")
    print("Instálelas en el entorno analitica con: conda install pandas openpyxl")
    raise SystemExit(1) from exc


# =============================================================================
# CONFIGURACIÓN
# =============================================================================
ROOT = Path(r"G:\Mi unidad\8.UES\PAGINA INDICADORES\UESVALLE")
RAW_DIR = ROOT / "data" / "sedes_educativas" / "raw"
CURRENT_DIR = ROOT / "data" / "sedes_educativas" / "current"
DOCS_DIR = ROOT / "docs" / "sedes_educativas"
ARCHIVE_DIR = ROOT / "archive" / "sedes_educativas"

INPUT_IRCA = RAW_DIR / "Escuelas_IRCAS.xlsx"
INPUT_SEDES = RAW_DIR / "SedesUESVALLE.xlsx"
INPUT_MAESTRO_ACUEDUCTOS = RAW_DIR / "maestro_acueductos_uesvalle_2026_actualizado.csv"
INPUT_MEN_SEDES = RAW_DIR / "MEN_SEDES_EDUCATIVAS_PREESCOLAR_BASICA_Y_MEDIA_20260605.csv"

OUT_BASE_CSV = CURRENT_DIR / "sedes_educativas_irca_consolidado.csv"
OUT_METADATA = CURRENT_DIR / "metadata_sedes_educativas.json"
OUT_RESUMEN_MUNICIPIO = CURRENT_DIR / "resumen_sedes_municipio.csv"
OUT_RESUMEN_RIESGO = CURRENT_DIR / "resumen_sedes_riesgo.csv"
OUT_RESUMEN_ACUEDUCTO = CURRENT_DIR / "resumen_acueducto.csv"
OUT_CONTROL_CRUCE = CURRENT_DIR / "control_cruce_sedes_irca.csv"
OUT_SIN_GEO = CURRENT_DIR / "sedes_sin_georreferenciar.csv"
OUT_VALIDAR_FECHAS = CURRENT_DIR / "validar_fechas_sedes.csv"
OUT_ACUEDUCTOS_MAPA = CURRENT_DIR / "maestro_acueductos_mapa.csv"
OUT_VALIDACION_MATRICULA = CURRENT_DIR / "validacion_matricula_fuentes.csv"
OUT_VALIDACION_MEN = CURRENT_DIR / "validacion_men_sedes.csv"
OUT_XLSX = DOCS_DIR / "SEDES_EDUCATIVAS_IRCA_BASE_CONSOLIDADA_CONTROL.xlsx"

FECHA_CORTE = date(2026, 6, 4)

RISK_ORDER = ["SIN RIESGO", "BAJO", "MEDIO", "ALTO", "INVIABLE SANITARIAMENTE", "SIN DATO"]
RISK_FACTOR = {
    "SIN RIESGO": 0,
    "BAJO": 1,
    "MEDIO": 2,
    "ALTO": 3,
    "INVIABLE SANITARIAMENTE": 4,
    "SIN DATO": 0.5,
}


# =============================================================================
# FUNCIONES DE NORMALIZACIÓN
# =============================================================================
def clean(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_text(value) -> str:
    text = clean(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def normalize_code(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().replace(".0", "")
    text = re.sub(r"\D", "", text)
    return "" if text in ("", "0") else text


def normalize_municipality_key(value) -> str:
    """Normaliza municipio para cruces contra MEN."""
    name = normalize_text(value)
    aliases = {
        "GUADALAJARA DE BUGA": "BUGA",
        "EL AGUILA": "EL AGUILA",
        "EL CAIRO": "EL CAIRO",
        "EL CERRITO": "EL CERRITO",
        "LA CUMBRE": "LA CUMBRE",
        "LA UNION": "LA UNION",
        "LA VICTORIA": "LA VICTORIA",
    }
    return aliases.get(name, name)


def positive_or_none(value):
    parsed = numeric(value, None)
    if parsed is None or parsed <= 0:
        return None
    return int(parsed) if float(parsed).is_integer() else parsed


def coordinate_is_valid(lat, lon) -> bool:
    """Valida coordenadas aproximadas para Valle del Cauca."""
    if lat is None or lon is None:
        return False
    try:
        lat = float(lat)
        lon = float(lon)
    except (TypeError, ValueError):
        return False
    return lat != 0 and lon != 0 and 2.5 <= lat <= 5.5 and -78.5 <= lon <= -75.0



def numeric(value, default=0):
    if pd.isna(value) or clean(value) == "":
        return default
    if isinstance(value, (int, float)):
        return value
    text = clean(value).replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return default


def normalize_risk(value) -> str:
    text = normalize_text(value)
    mapping = {
        "SIN RIESGO": "SIN RIESGO",
        "RIESGO BAJO": "BAJO",
        "BAJO": "BAJO",
        "RIESGO MEDIO": "MEDIO",
        "MEDIO": "MEDIO",
        "RIESGO ALTO": "ALTO",
        "ALTO": "ALTO",
        "INVIABLE SANITARIAMENTE": "INVIABLE SANITARIAMENTE",
        "INVIABLE": "INVIABLE SANITARIAMENTE",
        "SD": "SIN DATO",
        "S/D": "SIN DATO",
        "SIN DATO": "SIN DATO",
        "": "SIN DATO",
    }
    return mapping.get(text, "SIN DATO")


def parse_date(value):
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return pd.NaT
    return parsed.normalize()


def score_priority(row) -> float:
    risk = row["NIVEL_RIESGO_IRCA"]
    factor = RISK_FACTOR.get(risk, 0.5)
    students = float(row.get("ESTUDIANTES_ACTUAL", 0) or 0)
    concept = normalize_text(row.get("CONCEPTO_SANITARIO", ""))
    concept_factor = 1.2 if "DESFAVORABLE" in concept else (1.05 if "REQUERIMIENTO" in concept else 1.0)
    return round(factor * math.log10(students + 1) * concept_factor, 2)


def priority_band(score: float) -> str:
    if score >= 8:
        return "MUY ALTA"
    if score >= 6:
        return "ALTA"
    if score >= 4:
        return "MEDIA"
    return "BAJA"


def ensure_inputs() -> None:
    for folder in (RAW_DIR, CURRENT_DIR, DOCS_DIR, ARCHIVE_DIR):
        folder.mkdir(parents=True, exist_ok=True)

    missing = [path for path in (INPUT_IRCA, INPUT_SEDES, INPUT_MAESTRO_ACUEDUCTOS, INPUT_MEN_SEDES) if not path.exists()]
    if missing:
        print("[ERROR] No se encontraron los archivos fuente:")
        for path in missing:
            print(f"  - {path}")
        print("\nCopie los dos Excel en la carpeta raw y ejecute nuevamente.")
        raise SystemExit(1)


def backup_existing(path: Path) -> None:
    if path.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_folder = ARCHIVE_DIR / timestamp
        backup_folder.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, backup_folder / path.name)


# =============================================================================
# PROCESAMIENTO
# =============================================================================

def load_men_sedes() -> tuple[dict, dict]:
    """
    Carga la fuente nacional MEN y construye índices por código DANE de sede
    y por municipio para búsqueda por nombre.
    """
    men = pd.read_csv(INPUT_MEN_SEDES, dtype=object, encoding="utf-8-sig", sep=None, engine="python")
    men.columns = [clean(column) for column in men.columns]
    men = men.loc[:, [c for c in men.columns if c]]

    if "DEPARTAMENTO" in men.columns:
        men = men[men["DEPARTAMENTO"].map(normalize_text).eq("VALLE DEL CAUCA")].copy()

    for column in ["CODIGO_DANE_SEDE", "MUNICIPIO", "NOMBRE_ESTABLECIMIENTO", "NOMBRE_SEDE"]:
        if column not in men.columns:
            raise ValueError(f"La fuente MEN no contiene la columna obligatoria: {column}")

    men["_COD_SEDE"] = men["CODIGO_DANE_SEDE"].map(normalize_code)
    men["_MUN_KEY"] = men["MUNICIPIO"].map(normalize_municipality_key)
    men["_NOMBRE_SEDE_NORM"] = men["NOMBRE_SEDE"].map(normalize_text)
    men["_NOMBRE_EST_NORM"] = men["NOMBRE_ESTABLECIMIENTO"].map(normalize_text)

    if "TOTAL_MATRICULA" in men.columns:
        men["_TOTAL_MATRICULA"] = men["TOTAL_MATRICULA"].map(lambda value: positive_or_none(value))
    else:
        men["_TOTAL_MATRICULA"] = None

    if "COORDENADA_Y_SEDE" in men.columns:
        men["_LAT_MEN"] = men["COORDENADA_Y_SEDE"].map(lambda value: numeric(value, None))
    else:
        men["_LAT_MEN"] = None

    if "COORDENADA_X_SEDE" in men.columns:
        men["_LON_MEN"] = men["COORDENADA_X_SEDE"].map(lambda value: numeric(value, None))
    else:
        men["_LON_MEN"] = None

    men["_COORD_MEN_OK"] = men.apply(lambda row: coordinate_is_valid(row["_LAT_MEN"], row["_LON_MEN"]), axis=1)

    by_code = {}
    by_municipality = defaultdict(list)

    for _, row in men.iterrows():
        code = row["_COD_SEDE"]
        if code and code not in by_code:
            by_code[code] = row
        by_municipality[row["_MUN_KEY"]].append(row)

    return by_code, by_municipality


def find_men_match(irca_row: pd.Series, men_by_code: dict, men_by_municipality: dict):
    """
    Busca coincidencia en MEN. Prioriza código DANE de sede.
    Para registros sin código o no encontrados, usa coincidencia fuerte por nombre dentro del municipio.
    """
    code = irca_row.get("_COD", "")
    if code and code in men_by_code:
        return men_by_code[code], "CODIGO_DANE_SEDE", 1.0

    municipality = normalize_municipality_key(irca_row.get("MUNICIPIO"))
    candidates = men_by_municipality.get(municipality, [])
    if not candidates:
        return None, "SIN_CANDIDATOS_MUNICIPIO", None

    source_name = normalize_text(irca_row.get("NOMBRE_INS_PAE"))
    scored = []

    for candidate in candidates:
        men_sede = candidate.get("_NOMBRE_SEDE_NORM", "")
        men_est = candidate.get("_NOMBRE_EST_NORM", "")
        score_name = SequenceMatcher(None, source_name, men_sede).ratio() if men_sede else 0
        score_full = SequenceMatcher(None, source_name, f"{men_est} {men_sede}".strip()).ratio() if (men_est or men_sede) else 0
        score_contains = 0

        if men_sede and len(men_sede) >= 6 and men_sede in source_name:
            score_contains = 0.93
        if men_est and len(men_est) >= 8 and men_est in source_name and men_sede and men_sede in source_name:
            score_contains = 0.98

        score = max(score_name, score_full, score_contains)
        scored.append((score, candidate))

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best_candidate = scored[0]

    if best_score >= 0.90:
        if len(scored) > 1 and scored[1][0] >= best_score - 0.02:
            return None, "COINCIDENCIA_NOMBRE_FUERTE_CON_POSIBLE_EMPATE", round(best_score, 3)
        return best_candidate, "COINCIDENCIA_NOMBRE_FUERTE", round(best_score, 3)

    if best_score >= 0.82:
        return None, "POSIBLE_COINCIDENCIA_REVISAR", round(best_score, 3)

    return None, "SIN_COINCIDENCIA_FUERTE", round(best_score, 3)


def load_sources() -> tuple[pd.DataFrame, pd.DataFrame]:
    irca = pd.read_excel(INPUT_IRCA, dtype=object)
    sedes = pd.read_excel(INPUT_SEDES, dtype=object)

    irca.columns = [clean(column) for column in irca.columns]
    sedes.columns = [clean(column) for column in sedes.columns]

    irca = irca.loc[:, [c for c in irca.columns if c]]
    sedes = sedes.loc[:, [c for c in sedes.columns if c]]

    irca["ID_SEDE_IRCA"] = range(1, len(irca) + 1)
    sedes["ID_SEDESUESVALLE"] = range(1, len(sedes) + 1)

    for df in (irca, sedes):
        df["_COD"] = df["COD_DANE"].map(normalize_code)
        df["_MUN"] = df["MUNICIPIO"].map(normalize_text)
        df["_NOMBRE"] = df["NOMBRE_INS_PAE"].map(normalize_text)
        df["_KEY_NOMBRE"] = df["_MUN"] + "|" + df["_NOMBRE"]

    sedes["_LON"] = pd.to_numeric(sedes.get("LONGITUD"), errors="coerce")
    sedes["_LAT"] = pd.to_numeric(sedes.get("LATITUD"), errors="coerce")
    sedes["_COORD_OK"] = sedes["_LON"].notna() & sedes["_LAT"].notna()

    return irca, sedes


def choose_match(irca_row: pd.Series, sedes: pd.DataFrame):
    code = irca_row["_COD"]
    name_key = irca_row["_KEY_NOMBRE"]

    candidates_code = sedes[sedes["_COD"].eq(code)] if code else sedes.iloc[0:0]
    candidates_name = sedes[sedes["_KEY_NOMBRE"].eq(name_key)] if name_key != "|" else sedes.iloc[0:0]
    exact_code_name = candidates_code[candidates_code["_KEY_NOMBRE"].eq(name_key)]

    selected = None
    method = "SIN_COINCIDENCIA"
    status = "PENDIENTE"
    observation = ""

    if len(candidates_code) == 1:
        selected = candidates_code.iloc[0]
        method = "COD_DANE_UNICO"
        status = "CRUCE_OK"
    elif len(exact_code_name) == 1:
        selected = exact_code_name.iloc[0]
        method = "COD_DANE_DUPLICADO+NOMBRE_EXACTO"
        status = "CRUCE_OK_CONTROLADO"
    elif len(candidates_code) > 1:
        coords = candidates_code.loc[candidates_code["_COORD_OK"], ["_LAT", "_LON"]].drop_duplicates()
        if len(coords) == 1:
            selected = candidates_code[candidates_code["_COORD_OK"]].iloc[0]
            method = "COD_DANE_DUPLICADO_COORDENADA_COMUN"
            status = "VALIDAR_DUPLICADO"
            observation = "Código DANE duplicado en SedesUESVALLE; se asignó coordenada común."
        else:
            method = "COD_DANE_DUPLICADO_SIN_RESOLVER"
            status = "VALIDAR_DUPLICADO"
            observation = "Código DANE con múltiples candidatos; requiere selección manual."
    elif len(candidates_name) == 1:
        selected = candidates_name.iloc[0]
        method = "MUNICIPIO+NOMBRE_EXACTO"
        status = "CRUCE_RECUPERADO"
    elif len(candidates_name) > 1:
        coords = candidates_name.loc[candidates_name["_COORD_OK"], ["_LAT", "_LON"]].drop_duplicates()
        if len(coords) == 1:
            selected = candidates_name[candidates_name["_COORD_OK"]].iloc[0]
            method = "NOMBRE_DUPLICADO_COORDENADA_COMUN"
            status = "VALIDAR_DUPLICADO"
            observation = "Nombre duplicado; se asignó coordenada común."
        else:
            method = "NOMBRE_DUPLICADO_SIN_RESOLVER"
            status = "VALIDAR_DUPLICADO"
            observation = "Nombre exacto con múltiples candidatos; requiere selección manual."

    return selected, method, status, observation, len(candidates_code), len(candidates_name)


def consolidate(irca: pd.DataFrame, sedes: pd.DataFrame, men_by_code: dict, men_by_municipality: dict) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    records = []
    controls = []
    validation_men = []
    validation_matricula = []

    for _, row in irca.iterrows():
        matched, method, status, observation, n_code, n_name = choose_match(row, sedes)
        men_match, men_method, men_score = find_men_match(row, men_by_code, men_by_municipality)

        has_match = matched is not None

        lat_sedes = numeric(matched.get("LATITUD"), None) if has_match else None
        lon_sedes = numeric(matched.get("LONGITUD"), None) if has_match else None
        coord_sedes_ok = coordinate_is_valid(lat_sedes, lon_sedes)

        lat_men = numeric(men_match.get("_LAT_MEN"), None) if men_match is not None else None
        lon_men = numeric(men_match.get("_LON_MEN"), None) if men_match is not None else None
        coord_men_ok = bool(men_match.get("_COORD_MEN_OK")) if men_match is not None else False

        if coord_sedes_ok:
            lat_final = lat_sedes
            lon_final = lon_sedes
            fuente_coord = "SEDESUESVALLE"
            estado_coord = "COORDENADA_ORIGINAL_SEDESUESVALLE"
        elif coord_men_ok:
            lat_final = lat_men
            lon_final = lon_men
            fuente_coord = "MEN_SEDES"
            estado_coord = "COMPLEMENTADA_CON_MEN"
        else:
            lat_final = None
            lon_final = None
            fuente_coord = "SIN_COORDENADA"
            estado_coord = "SIN_COORDENADA"

        estudiantes_irca = positive_or_none(row.get("Número de Estudiantes_Actual"))
        estudiantes_sedes = positive_or_none(matched.get("Total_general_INS")) if has_match else None
        estudiantes_men = positive_or_none(men_match.get("_TOTAL_MATRICULA")) if men_match is not None else None

        if estudiantes_irca is not None:
            estudiantes_final = estudiantes_irca
            fuente_matricula = "ESCUELAS_IRCAS"
            estado_matricula = "DATO_PRINCIPAL_ESCUELAS_IRCAS"
            requiere_validacion_matricula = "NO"
        elif estudiantes_sedes is not None:
            estudiantes_final = estudiantes_sedes
            fuente_matricula = "SEDESUESVALLE"
            estado_matricula = "COMPLEMENTADA_CON_SEDESUESVALLE"
            requiere_validacion_matricula = "NO"
        elif estudiantes_men is not None and men_method in ("CODIGO_DANE_SEDE", "COINCIDENCIA_NOMBRE_FUERTE"):
            estudiantes_final = estudiantes_men
            fuente_matricula = "MEN_SEDES_2019"
            estado_matricula = "COMPLEMENTADA_CON_MEN_REFERENCIA_HISTORICA"
            requiere_validacion_matricula = "SI"
        else:
            estudiantes_final = 0
            fuente_matricula = "SIN_FUENTE"
            estado_matricula = "SIN_DATO_MATRICULA"
            requiere_validacion_matricula = "SI"

        out = {
            "ID_SEDE_IRCA": row["ID_SEDE_IRCA"],
            "COD_DANE": row["_COD"],
            "MUNICIPIO": clean(row.get("MUNICIPIO")),
            "NOMBRE_SEDE_IRCA": clean(row.get("NOMBRE_INS_PAE")),
            "ZONA_IRCA": normalize_text(row.get("ZONA")),
            "DIRECCION_IRCA": clean(row.get("DIRECCION_PAE")),
            "NOMBRE_ABREVIADO_IRCA": clean(row.get("NOMBREABREVIADO")),
            "TELEFONO_IRCA": clean(row.get("TELEFONO")),

            "ESTUDIANTES_ACTUAL": int(estudiantes_final or 0),
            "ESTUDIANTES_IRCAS": int(estudiantes_irca or 0),
            "ESTUDIANTES_SEDESUESVALLE": int(estudiantes_sedes or 0),
            "ESTUDIANTES_MEN_REFERENCIA": int(estudiantes_men or 0),
            "FUENTE_MATRICULA_FINAL": fuente_matricula,
            "ESTADO_COMPLEMENTO_MATRICULA": estado_matricula,
            "REQUIERE_VALIDACION_MATRICULA": requiere_validacion_matricula,

            "CONECTADA_ACUEDUCTO": normalize_text(row.get("Conectada_Acueducto (Si/No)")),
            "NOMBRE_ACUEDUCTO": clean(row.get("NOMBRE DE ACUEDUCTO")),
            "IRCA_ORIGINAL": clean(row.get("IRCA")),
            "NIVEL_RIESGO_IRCA": normalize_risk(row.get("IRCA")),
            "CODIGO_SISA_ACUEDUCTO": clean(row.get("codigo SISA de acueducto veredal")),
            "TRATAMIENTO_EN_ESCUELA": clean(row.get("TIPO DE TRATAMIENTO EN ESCUELA (SI EXISTE)\nFILTRO, PLANTA COMPACTA, ACUEDUCTO PROPIO")),
            "FUENTE_ABASTO": clean(row.get("FUENTE DE ABASTO(RIO, QUEBRADA, LLUVIA…)")),
            "CALIFICACION_BLOQUE_AGUA": clean(row.get("Calificacion Bloque Agua y Sanemiento (acta IVC)")),
            "CONCEPTO_SANITARIO": clean(row.get("CONCEPTO SANITARIO")),
            "FECHA_VISITA": parse_date(row.get("Fecha de visita")),

            "TIPO_CRUCE": method,
            "ESTADO_CRUCE": status,
            "ESTADO_GEORREFERENCIACION": "GEORREFERENCIADA" if estado_coord != "SIN_COORDENADA" else "SIN_COORDENADAS",
            "OBSERVACION_CRUCE": observation,
            "ID_SEDESUESVALLE": matched["ID_SEDESUESVALLE"] if has_match else "",

            "INSTITUCION_INS": clean(matched.get("INSTITUCION_INS")) if has_match else "",
            "SEDE_INS": clean(matched.get("SEDE_INS")) if has_match else "",
            "ESTADO_SEDE": clean(matched.get("ESTADO")) if has_match else "",
            "NATURALEZA": clean(matched.get("NATURALEZA")) if has_match else "",
            "JORNADA": clean(matched.get("JORNADA")) if has_match else "",
            "NIVEL_EDUCATIVO": clean(matched.get("NIVEL")) if has_match else "",
            "DIRECCION_MAPA": clean(matched.get("DIRECCION_MAPA")) if has_match else "",

            "LONGITUD": lon_final,
            "LATITUD": lat_final,
            "LONGITUD_SEDESUESVALLE": lon_sedes,
            "LATITUD_SEDESUESVALLE": lat_sedes,
            "LONGITUD_MEN": lon_men,
            "LATITUD_MEN": lat_men,
            "FUENTE_COORDENADA": fuente_coord,
            "ESTADO_COORDENADA": estado_coord,

            "BENEFICIARIOS_PAE_2025_1SI": int(numeric(matched.get("Beneficiarios_PAE_2025_1SI"), 0)) if has_match else 0,
            "TOTAL_GENERAL_INS": int(estudiantes_sedes or 0),

            "MEN_ANIO": clean(men_match.get("AÑO")) if men_match is not None else "",
            "MEN_ESTABLECIMIENTO": clean(men_match.get("NOMBRE_ESTABLECIMIENTO")) if men_match is not None else "",
            "MEN_SEDE": clean(men_match.get("NOMBRE_SEDE")) if men_match is not None else "",
            "MEN_ZONA": clean(men_match.get("ZONA")) if men_match is not None else "",
            "MEN_DIRECCION": clean(men_match.get("DIRECCION")) if men_match is not None else "",
            "MEN_METODO_CRUCE": men_method,
            "MEN_SCORE_CRUCE": men_score if men_score is not None else "",
        }

        out["SCORE_PRIORIDAD"] = score_priority(out)
        out["NIVEL_PRIORIDAD"] = priority_band(out["SCORE_PRIORIDAD"])
        records.append(out)

        controls.append({
            "ID_SEDE_IRCA": row["ID_SEDE_IRCA"],
            "COD_DANE": row["_COD"],
            "MUNICIPIO": clean(row.get("MUNICIPIO")),
            "NOMBRE_SEDE_IRCA": clean(row.get("NOMBRE_INS_PAE")),
            "CANDIDATOS_POR_CODIGO": n_code,
            "CANDIDATOS_POR_NOMBRE": n_name,
            "TIPO_CRUCE": method,
            "ESTADO_CRUCE": status,
            "ESTADO_GEORREFERENCIACION": out["ESTADO_GEORREFERENCIACION"],
            "OBSERVACION": observation,
        })

        validation_men.append({
            "ID_SEDE_IRCA": row["ID_SEDE_IRCA"],
            "COD_DANE": row["_COD"],
            "MUNICIPIO_BASE": clean(row.get("MUNICIPIO")),
            "NOMBRE_SEDE_BASE": clean(row.get("NOMBRE_INS_PAE")),
            "MEN_COINCIDE": "SI" if men_match is not None else "NO",
            "MEN_METODO_CRUCE": men_method,
            "MEN_SCORE_CRUCE": men_score if men_score is not None else "",
            "MEN_ANIO": clean(men_match.get("AÑO")) if men_match is not None else "",
            "MEN_ESTABLECIMIENTO": clean(men_match.get("NOMBRE_ESTABLECIMIENTO")) if men_match is not None else "",
            "MEN_SEDE": clean(men_match.get("NOMBRE_SEDE")) if men_match is not None else "",
            "MEN_TOTAL_MATRICULA": int(estudiantes_men or 0),
            "MEN_LATITUD": lat_men,
            "MEN_LONGITUD": lon_men,
            "MEN_COORDENADA_VALIDA": "SI" if coord_men_ok else "NO",
            "USO_COORDENADA_MEN": "SI" if fuente_coord == "MEN_SEDES" else "NO",
            "USO_MATRICULA_MEN": "SI" if fuente_matricula == "MEN_SEDES_2019" else "NO",
        })

        validation_matricula.append({
            "ID_SEDE_IRCA": row["ID_SEDE_IRCA"],
            "COD_DANE": row["_COD"],
            "MUNICIPIO": clean(row.get("MUNICIPIO")),
            "NOMBRE_SEDE_IRCA": clean(row.get("NOMBRE_INS_PAE")),
            "NIVEL_RIESGO_IRCA": out["NIVEL_RIESGO_IRCA"],
            "ESTUDIANTES_IRCAS": out["ESTUDIANTES_IRCAS"],
            "ESTUDIANTES_SEDESUESVALLE": out["ESTUDIANTES_SEDESUESVALLE"],
            "ESTUDIANTES_MEN_REFERENCIA": out["ESTUDIANTES_MEN_REFERENCIA"],
            "ESTUDIANTES_ACTUAL_FINAL": out["ESTUDIANTES_ACTUAL"],
            "FUENTE_MATRICULA_FINAL": fuente_matricula,
            "ESTADO_COMPLEMENTO_MATRICULA": estado_matricula,
            "REQUIERE_VALIDACION_MATRICULA": requiere_validacion_matricula,
            "MEN_METODO_CRUCE": men_method,
            "MEN_SCORE_CRUCE": men_score if men_score is not None else "",
        })

    return pd.DataFrame(records), pd.DataFrame(controls), pd.DataFrame(validation_men), pd.DataFrame(validation_matricula)


def create_aqueduct_map_layer() -> pd.DataFrame:
    """Prepara la capa de puntos de acueductos que será consumida por el HTML."""
    maestro = pd.read_csv(INPUT_MAESTRO_ACUEDUCTOS, encoding="utf-8-sig", sep=None, engine="python")
    maestro.columns = [clean(column) for column in maestro.columns]

    required = ["LATITUD", "LONGITUD", "CODIGO_SISTEMA", "MUNICIPIO", "PERSONA_PRESTADORA"]
    missing = [column for column in required if column not in maestro.columns]
    if missing:
        raise ValueError("El maestro de acueductos no contiene las columnas obligatorias: " + ", ".join(missing))

    for column in ["LATITUD", "LONGITUD", "POBLACION", "SUSCRIPTORES"]:
        if column in maestro.columns:
            maestro[column] = pd.to_numeric(maestro[column], errors="coerce")

    columns = [
        "ID_MAESTRO", "CODIGO_SISTEMA", "MUNICIPIO", "AMBITO", "TIPO_SISTEMA", "ARO",
        "PERSONA_PRESTADORA", "LOCALIDADES_ABASTECIDAS", "SUSCRIPTORES", "POBLACION",
        "TIPO_SISTEMA_TRATAMIENTO", "ALGUN_TIPO_TRATAMIENTO", "ESTADO_MAESTRO",
        "LATITUD", "LONGITUD", "VALIDACION_SISTEMA"
    ]
    available = [column for column in columns if column in maestro.columns]
    capa = maestro.loc[maestro["LATITUD"].notna() & maestro["LONGITUD"].notna(), available].copy()
    capa["MUNICIPIO"] = capa["MUNICIPIO"].map(normalize_text)
    return capa


def create_summaries(base: pd.DataFrame):
    risk = (
        base.groupby("NIVEL_RIESGO_IRCA", dropna=False)
        .agg(SEDES=("ID_SEDE_IRCA", "count"), ESTUDIANTES=("ESTUDIANTES_ACTUAL", "sum"))
        .reindex(RISK_ORDER, fill_value=0)
        .reset_index()
    )

    base_calc = base.copy()
    base_calc["SEDE_ALTO_INVIABLE"] = base_calc["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]).astype(int)
    base_calc["EST_MEDIO_MAS"] = base_calc["ESTUDIANTES_ACTUAL"].where(
        base_calc["NIVEL_RIESGO_IRCA"].isin(["MEDIO", "ALTO", "INVIABLE SANITARIAMENTE"]), 0
    )
    base_calc["EST_ALTO_INVIABLE"] = base_calc["ESTUDIANTES_ACTUAL"].where(
        base_calc["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]), 0
    )
    base_calc["GEORREFERENCIADA"] = base_calc["ESTADO_GEORREFERENCIACION"].eq("GEORREFERENCIADA").astype(int)

    counts = pd.crosstab(base_calc["MUNICIPIO"], base_calc["NIVEL_RIESGO_IRCA"])
    counts = counts.reindex(columns=RISK_ORDER, fill_value=0).reset_index()
    totals = (
        base_calc.groupby("MUNICIPIO", as_index=False)
        .agg(
            SEDES_TOTAL=("ID_SEDE_IRCA", "count"),
            ESTUDIANTES_TOTAL=("ESTUDIANTES_ACTUAL", "sum"),
            SEDES_ALTO_INVIABLE=("SEDE_ALTO_INVIABLE", "sum"),
            ESTUDIANTES_MEDIO_MAS=("EST_MEDIO_MAS", "sum"),
            ESTUDIANTES_ALTO_INVIABLE=("EST_ALTO_INVIABLE", "sum"),
            SEDES_GEORREFERENCIADAS=("GEORREFERENCIADA", "sum"),
        )
    )
    municipality = totals.merge(counts, on="MUNICIPIO", how="left")
    municipality = municipality.sort_values(["SEDES_ALTO_INVIABLE", "ESTUDIANTES_ALTO_INVIABLE"], ascending=False)

    aqueduct = (
        base_calc.assign(
            CODIGO_SISA_ACUEDUCTO=base_calc["CODIGO_SISA_ACUEDUCTO"].replace("", "SIN CODIGO SISA"),
            NOMBRE_ACUEDUCTO=base_calc["NOMBRE_ACUEDUCTO"].replace("", "SIN NOMBRE DE ACUEDUCTO"),
        )
        .groupby(["CODIGO_SISA_ACUEDUCTO", "NOMBRE_ACUEDUCTO"], dropna=False, as_index=False)
        .agg(
            MUNICIPIOS=("MUNICIPIO", lambda s: ", ".join(sorted(set(x for x in s if x)))),
            SEDES_TOTAL=("ID_SEDE_IRCA", "count"),
            ESTUDIANTES_TOTAL=("ESTUDIANTES_ACTUAL", "sum"),
            SEDES_ALTO_INVIABLE=("SEDE_ALTO_INVIABLE", "sum"),
            ESTUDIANTES_ALTO_INVIABLE=("EST_ALTO_INVIABLE", "sum"),
        )
        .sort_values(["SEDES_ALTO_INVIABLE", "ESTUDIANTES_ALTO_INVIABLE"], ascending=False)
    )
    return risk, municipality, aqueduct


# =============================================================================
# EXCEL DE CONTROL
# =============================================================================
def build_excel(base, control, risk, municipality, aqueduct, without_geo, validate_dates, validation_men, validation_matricula) -> None:
    for path in (OUT_XLSX,):
        backup_existing(path)

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        # Resumen inicial; se complementa con fórmulas y formato después.
        resumen_cruce = (
            base.groupby("ESTADO_CRUCE", as_index=False)
            .agg(REGISTROS=("ID_SEDE_IRCA", "count"))
            .sort_values("REGISTROS", ascending=False)
        )
        risk.to_excel(writer, sheet_name="RESUMEN_EJECUTIVO", index=False, startrow=14, startcol=0)
        resumen_cruce.to_excel(writer, sheet_name="RESUMEN_EJECUTIVO", index=False, startrow=14, startcol=4)
        municipality.head(10).to_excel(writer, sheet_name="RESUMEN_EJECUTIVO", index=False, startrow=24, startcol=0)

        base.to_excel(writer, sheet_name="BASE_CONSOLIDADA", index=False, startrow=2)
        control.to_excel(writer, sheet_name="CONTROL_CRUCE", index=False, startrow=2)
        base[base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"])].sort_values(
            ["SCORE_PRIORIDAD", "ESTUDIANTES_ACTUAL"], ascending=False
        ).to_excel(writer, sheet_name="SEDES_CRITICAS", index=False, startrow=2)
        without_geo.to_excel(writer, sheet_name="SEDES_SIN_COORDENADAS", index=False, startrow=2)
        municipality.to_excel(writer, sheet_name="RESUMEN_MUNICIPIO", index=False, startrow=2)
        aqueduct.to_excel(writer, sheet_name="RESUMEN_ACUEDUCTO", index=False, startrow=2)
        validate_dates.to_excel(writer, sheet_name="VALIDAR_FECHAS", index=False, startrow=2)
        validation_men.to_excel(writer, sheet_name="VALIDACION_MEN", index=False, startrow=2)
        validation_matricula.to_excel(writer, sheet_name="VALIDACION_MATRICULA", index=False, startrow=2)

        rules = pd.DataFrame([
            ["Fuente", "Escuelas_IRCAS.xlsx", "Universo sanitario; cada registro se conserva.", "Base principal"],
            ["Fuente", "SedesUESVALLE.xlsx", "Coordenadas y atributos educativos complementarios.", "Cruce espacial"],
            ["Cruce", "COD_DANE_UNICO", "Coincidencia única por código DANE.", "Automático"],
            ["Cruce", "COD_DANE_DUPLICADO+NOMBRE_EXACTO", "Duplicado resuelto por nombre normalizado.", "Controlado"],
            ["Cruce", "MUNICIPIO+NOMBRE_EXACTO", "Recuperado cuando no se resuelve por código.", "Recuperado"],
            ["IRCA", "Normalización", "RIESGO ALTO/ALTO → ALTO; SD → SIN DATO.", "Simbología"],
            ["Prioridad", "SCORE_PRIORIDAD", "Factor de riesgo × log10(estudiantes + 1) × factor sanitario.", "Ranking"],
            ["Fecha", "Fecha de corte", f"Validar fechas posteriores a {FECHA_CORTE.isoformat()}.", "Control"],
        ], columns=["SECCION", "CAMPO_REGLA", "DESCRIPCION", "USO"])
        rules.to_excel(writer, sheet_name="DICCIONARIO_REGLAS", index=False, startrow=2)

    wb = load_workbook(OUT_XLSX)
    blue = "2E64D2"
    light_blue = "EAF1FF"
    dark = "14213D"
    border = Side(style="thin", color="D9E3F0")
    title_fill = PatternFill("solid", fgColor=blue)
    header_fill = PatternFill("solid", fgColor=blue)

    titles = {
        "RESUMEN_EJECUTIVO": "Tablero Sedes Educativas – Base consolidada de calidad del agua",
        "BASE_CONSOLIDADA": "Base consolidada sanitaria y espacial – una fila por sede evaluada",
        "CONTROL_CRUCE": "Control de cruce entre Escuelas_IRCAS y SedesUESVALLE",
        "SEDES_CRITICAS": "Sedes clasificadas en riesgo alto o inviable sanitariamente",
        "SEDES_SIN_COORDENADAS": "Sedes pendientes de georreferenciación para el mapa",
        "RESUMEN_MUNICIPIO": "Resumen municipal para KPIs y gráficos",
        "RESUMEN_ACUEDUCTO": "Acueductos abastecedores y sedes asociadas",
        "VALIDAR_FECHAS": "Fechas vacías o posteriores al corte",
        "VALIDACION_MEN": "Validación contra fuente nacional MEN",
        "VALIDACION_MATRICULA": "Trazabilidad de matrícula final recomendada",
        "DICCIONARIO_REGLAS": "Diccionario y reglas de consolidación",
    }

    for ws in wb.worksheets:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ws.max_column, 4))
        cell = ws.cell(1, 1)
        cell.value = titles.get(ws.title, ws.title)
        cell.fill = title_fill
        cell.font = Font(color="FFFFFF", bold=True, size=15)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A4"

        header_row = 15 if ws.title == "RESUMEN_EJECUTIVO" else 3
        for cell in ws[header_row]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=border, right=border, top=border, bottom=border)
        ws.row_dimensions[header_row].height = 30
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

        for col in range(1, ws.max_column + 1):
            values = [clean(ws.cell(r, col).value) for r in range(1, min(ws.max_row, 60) + 1)]
            width = min(max(max((len(v) for v in values), default=10) + 2, 12), 42)
            ws.column_dimensions[get_column_letter(col)].width = width

    summary = wb["RESUMEN_EJECUTIVO"]
    summary["A3"] = "Objetivo: integrar la clasificación sanitaria asociada al agua que abastece las sedes educativas con su ubicación geográfica y caracterización educativa."
    summary.merge_cells("A3:L4")
    summary["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    summary["A3"].fill = PatternFill("solid", fgColor="F8FAFC")
    summary["A3"].font = Font(color="475467", size=10)

    kpis = [
        ("A6", "Total sedes evaluadas", len(base)),
        ("D6", "Sedes georreferenciadas", int((base["ESTADO_GEORREFERENCIACION"] == "GEORREFERENCIADA").sum())),
        ("G6", "Estudiantes reportados", int(base["ESTUDIANTES_ACTUAL"].sum())),
        ("J6", "Sedes alto/inviable", int(base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]).sum())),
        ("A10", "Estudiantes alto/inviable", int(base.loc[base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]), "ESTUDIANTES_ACTUAL"].sum())),
        ("D10", "Sedes sin IRCA", int((base["NIVEL_RIESGO_IRCA"] == "SIN DATO").sum())),
        ("G10", "Sedes sin coordenadas", int((base["ESTADO_GEORREFERENCIACION"] == "SIN_COORDENADAS").sum())),
        ("J10", "Cruces por validar/pendientes", int(base["ESTADO_CRUCE"].isin(["VALIDAR_DUPLICADO", "PENDIENTE"]).sum())),
    ]
    for start, label, value in kpis:
        col = summary[start].column
        row = summary[start].row
        summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        summary.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 2)
        summary.cell(row, col).value = label
        summary.cell(row, col).fill = PatternFill("solid", fgColor=light_blue)
        summary.cell(row, col).font = Font(bold=True, color=dark, size=10)
        summary.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        summary.cell(row + 1, col).value = value
        summary.cell(row + 1, col).font = Font(bold=True, color=dark, size=18)
        summary.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")
        summary.cell(row + 1, col).number_format = "#,##0"

    chart = BarChart()
    chart.title = "Sedes educativas por nivel de riesgo"
    chart.y_axis.title = "Sedes"
    chart.x_axis.title = "Nivel de riesgo"
    data = Reference(summary, min_col=2, min_row=15, max_row=21)
    categories = Reference(summary, min_col=1, min_row=16, max_row=21)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    summary.add_chart(chart, "E32")

    # Tablas estructuradas en hojas de datos
    for ws in wb.worksheets:
        if ws.title == "RESUMEN_EJECUTIVO":
            continue
        ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
        table_name = re.sub(r"[^A-Za-z0-9]", "", ws.title.title()) + "Tabla"
        table = Table(displayName=table_name[:25], ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    wb.save(OUT_XLSX)


def main() -> None:
    print("=" * 88)
    print(" NORMALIZAR Y CONSOLIDAR SEDES EDUCATIVAS - IRCA | UESVALLE")
    print("=" * 88)
    print(f"Repositorio: {ROOT}")
    print(f"Fuentes    : {RAW_DIR}")
    print(f"Salidas    : {CURRENT_DIR}")
    print()

    ensure_inputs()
    irca, sedes = load_sources()
    men_by_code, men_by_municipality = load_men_sedes()
    base, control, validation_men, validation_matricula = consolidate(irca, sedes, men_by_code, men_by_municipality)
    risk, municipality, aqueduct = create_summaries(base)
    aqueduct_map = create_aqueduct_map_layer()

    without_geo = base[base["ESTADO_GEORREFERENCIACION"] == "SIN_COORDENADAS"].copy()
    validate_dates = base[
        base["FECHA_VISITA"].isna()
        | (base["FECHA_VISITA"].dt.date > FECHA_CORTE)
    ][["ID_SEDE_IRCA", "COD_DANE", "MUNICIPIO", "NOMBRE_SEDE_IRCA", "FECHA_VISITA"]].copy()
    validate_dates["VALIDACION"] = validate_dates["FECHA_VISITA"].apply(
        lambda x: "FECHA VACÍA O NO INTERPRETABLE" if pd.isna(x) else "FECHA POSTERIOR AL CORTE"
    )

    for output in (OUT_BASE_CSV, OUT_METADATA, OUT_RESUMEN_MUNICIPIO, OUT_RESUMEN_RIESGO, OUT_RESUMEN_ACUEDUCTO, OUT_CONTROL_CRUCE, OUT_SIN_GEO, OUT_VALIDAR_FECHAS, OUT_ACUEDUCTOS_MAPA, OUT_VALIDACION_MEN, OUT_VALIDACION_MATRICULA):
        backup_existing(output)

    base.to_csv(OUT_BASE_CSV, index=False, encoding="utf-8-sig", sep=";")
    risk.to_csv(OUT_RESUMEN_RIESGO, index=False, encoding="utf-8-sig", sep=";")
    municipality.to_csv(OUT_RESUMEN_MUNICIPIO, index=False, encoding="utf-8-sig", sep=";")
    aqueduct.to_csv(OUT_RESUMEN_ACUEDUCTO, index=False, encoding="utf-8-sig", sep=";")
    control.to_csv(OUT_CONTROL_CRUCE, index=False, encoding="utf-8-sig", sep=";")
    without_geo.to_csv(OUT_SIN_GEO, index=False, encoding="utf-8-sig", sep=";")
    validate_dates.to_csv(OUT_VALIDAR_FECHAS, index=False, encoding="utf-8-sig", sep=";")
    aqueduct_map.to_csv(OUT_ACUEDUCTOS_MAPA, index=False, encoding="utf-8-sig", sep=";")
    validation_men.to_csv(OUT_VALIDACION_MEN, index=False, encoding="utf-8-sig", sep=";")
    validation_matricula.to_csv(OUT_VALIDACION_MATRICULA, index=False, encoding="utf-8-sig", sep=";")

    metadata = {
        "fecha_generacion": datetime.now().isoformat(timespec="seconds"),
        "fecha_corte_validacion": FECHA_CORTE.isoformat(),
        "fuentes": [INPUT_IRCA.name, INPUT_SEDES.name, INPUT_MAESTRO_ACUEDUCTOS.name, INPUT_MEN_SEDES.name],
        "acueductos_georreferenciados": int(len(aqueduct_map)),
        "sedes_coordenadas_complementadas_men": int((base["FUENTE_COORDENADA"] == "MEN_SEDES").sum()),
        "sedes_matricula_complementada_sedesuesvalle": int((base["FUENTE_MATRICULA_FINAL"] == "SEDESUESVALLE").sum()),
        "sedes_matricula_complementada_men": int((base["FUENTE_MATRICULA_FINAL"] == "MEN_SEDES_2019").sum()),
        "registros_sedes_irca": int(len(base)),
        "sedes_georreferenciadas": int((base["ESTADO_GEORREFERENCIACION"] == "GEORREFERENCIADA").sum()),
        "sedes_sin_coordenadas": int((base["ESTADO_GEORREFERENCIACION"] == "SIN_COORDENADAS").sum()),
        "sedes_alto_inviable": int(base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]).sum()),
        "estudiantes_alto_inviable": int(base.loc[base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]), "ESTUDIANTES_ACTUAL"].sum()),
        "riesgo_sedes": {risk_value: int((base["NIVEL_RIESGO_IRCA"] == risk_value).sum()) for risk_value in RISK_ORDER},
        "riesgo_estudiantes": {risk_value: int(base.loc[base["NIVEL_RIESGO_IRCA"] == risk_value, "ESTUDIANTES_ACTUAL"].sum()) for risk_value in RISK_ORDER},
        "estado_cruce": {str(key): int(value) for key, value in Counter(base["ESTADO_CRUCE"]).items()},
    }
    with OUT_METADATA.open("w", encoding="utf-8") as file:
        json.dump(metadata, file, ensure_ascii=False, indent=2)

    build_excel(base, control, risk, municipality, aqueduct, without_geo, validate_dates, validation_men, validation_matricula)

    print("[OK] Proceso finalizado.")
    print(f"  Universo sedes                 : {len(base):,}".replace(",", "."))
    print(f"  Sedes georreferenciadas        : {(base['ESTADO_GEORREFERENCIACION'] == 'GEORREFERENCIADA').sum():,}".replace(",", "."))
    print(f"  Sedes sin coordenadas          : {(base['ESTADO_GEORREFERENCIACION'] == 'SIN_COORDENADAS').sum():,}".replace(",", "."))
    print(f"  Sedes alto o inviable          : {base['NIVEL_RIESGO_IRCA'].isin(['ALTO', 'INVIABLE SANITARIAMENTE']).sum():,}".replace(",", "."))
    print(f"  Estudiantes alto o inviable    : {base.loc[base['NIVEL_RIESGO_IRCA'].isin(['ALTO', 'INVIABLE SANITARIAMENTE']), 'ESTUDIANTES_ACTUAL'].sum():,}".replace(",", "."))
    print(f"  Cruces validar/pendientes      : {base['ESTADO_CRUCE'].isin(['VALIDAR_DUPLICADO', 'PENDIENTE']).sum():,}".replace(",", "."))
    print(f"  Acueductos georreferenciados   : {len(aqueduct_map):,}".replace(",", "."))
    print(f"  Coordenadas complementadas MEN : {(base['FUENTE_COORDENADA'] == 'MEN_SEDES').sum():,}".replace(",", "."))
    print(f"  Matrícula complementada Sedes  : {(base['FUENTE_MATRICULA_FINAL'] == 'SEDESUESVALLE').sum():,}".replace(",", "."))
    print(f"  Matrícula complementada MEN    : {(base['FUENTE_MATRICULA_FINAL'] == 'MEN_SEDES_2019').sum():,}".replace(",", "."))
    print()
    print("Archivos generados:")
    print(f"  - {OUT_XLSX}")
    print(f"  - {OUT_BASE_CSV}")
    print(f"  - {OUT_METADATA}")
    print(f"  - {OUT_RESUMEN_MUNICIPIO}")
    print(f"  - {OUT_RESUMEN_RIESGO}")
    print(f"  - {OUT_RESUMEN_ACUEDUCTO}")
    print(f"  - {OUT_CONTROL_CRUCE}")
    print(f"  - {OUT_SIN_GEO}")
    print(f"  - {OUT_VALIDAR_FECHAS}")
    print(f"  - {OUT_ACUEDUCTOS_MAPA}")
    print(f"  - {OUT_VALIDACION_MEN}")
    print(f"  - {OUT_VALIDACION_MATRICULA}")


if __name__ == "__main__":
    main()
