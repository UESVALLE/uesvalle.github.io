# -*- coding: utf-8 -*-
"""
Normalización y consolidación del tablero Sedes Educativas - UESVALLE.

Entradas esperadas:
  data/sedes_educativas/raw/Escuelas_IRCAS.xlsx
  data/sedes_educativas/raw/SedesUESVALLE.xlsx

Salidas:
  data/sedes_educativas/current/sedes_educativas_irca_consolidado.csv
  data/sedes_educativas/current/metadata_sedes_educativas.json
  data/sedes_educativas/current/resumen_sedes_municipio.csv
  data/sedes_educativas/current/resumen_sedes_riesgo.csv
  data/sedes_educativas/current/resumen_acueducto.csv
  data/sedes_educativas/current/control_cruce_sedes_irca.csv
  data/sedes_educativas/current/sedes_sin_georreferenciar.csv
  data/sedes_educativas/current/maestro_acueductos_mapa.csv
  docs/sedes_educativas/SEDES_EDUCATIVAS_IRCA_BASE_CONSOLIDADA_CONTROL.xlsx

Versión: V1
Fecha: 2026-06-04
"""

from __future__ import annotations

import json
import math
import re
import shutil
import sys
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    print("[ERROR] Faltan librerías requeridas: pandas y openpyxl.")
    print("Instálelas en el entorno analitica con: conda install pandas openpyxl")
    raise SystemExit(1) from exc


# =============================================================================
# CONFIGURACIÓN
# =============================================================================
ROOT = Path(r"G:\Mi unidad\8.UES\PAGINA INDICADORES\UESVALLE")
RAW_DIR = ROOT / "data" / "sedes_educativas" / "raw"
CURRENT_DIR = ROOT / "data" / "sedes_educativas" / "current"
DOCS_DIR = ROOT / "docs" / "sedes_educativas"
ARCHIVE_DIR = ROOT / "archive" / "sedes_educativas"

INPUT_IRCA = RAW_DIR / "Escuelas_IRCAS.xlsx"
INPUT_SEDES = RAW_DIR / "SedesUESVALLE.xlsx"
INPUT_MAESTRO_ACUEDUCTOS = RAW_DIR / "maestro_acueductos_uesvalle_2026_actualizado.csv"

OUT_BASE_CSV = CURRENT_DIR / "sedes_educativas_irca_consolidado.csv"
OUT_METADATA = CURRENT_DIR / "metadata_sedes_educativas.json"
OUT_RESUMEN_MUNICIPIO = CURRENT_DIR / "resumen_sedes_municipio.csv"
OUT_RESUMEN_RIESGO = CURRENT_DIR / "resumen_sedes_riesgo.csv"
OUT_RESUMEN_ACUEDUCTO = CURRENT_DIR / "resumen_acueducto.csv"
OUT_CONTROL_CRUCE = CURRENT_DIR / "control_cruce_sedes_irca.csv"
OUT_SIN_GEO = CURRENT_DIR / "sedes_sin_georreferenciar.csv"
OUT_VALIDAR_FECHAS = CURRENT_DIR / "validar_fechas_sedes.csv"
OUT_ACUEDUCTOS_MAPA = CURRENT_DIR / "maestro_acueductos_mapa.csv"
OUT_XLSX = DOCS_DIR / "SEDES_EDUCATIVAS_IRCA_BASE_CONSOLIDADA_CONTROL.xlsx"

FECHA_CORTE = date(2026, 6, 4)

RISK_ORDER = ["SIN RIESGO", "BAJO", "MEDIO", "ALTO", "INVIABLE SANITARIAMENTE", "SIN DATO"]
RISK_FACTOR = {
    "SIN RIESGO": 0,
    "BAJO": 1,
    "MEDIO": 2,
    "ALTO": 3,
    "INVIABLE SANITARIAMENTE": 4,
    "SIN DATO": 0.5,
}


# =============================================================================
# FUNCIONES DE NORMALIZACIÓN
# =============================================================================
def clean(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_text(value) -> str:
    text = clean(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"\s+", " ", text).strip().upper()
    return text


def normalize_code(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().replace(".0", "")
    text = re.sub(r"\D", "", text)
    return "" if text in ("", "0") else text


def numeric(value, default=0):
    if pd.isna(value) or clean(value) == "":
        return default
    if isinstance(value, (int, float)):
        return value
    text = clean(value).replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return default


def normalize_risk(value) -> str:
    text = normalize_text(value)
    mapping = {
        "SIN RIESGO": "SIN RIESGO",
        "RIESGO BAJO": "BAJO",
        "BAJO": "BAJO",
        "RIESGO MEDIO": "MEDIO",
        "MEDIO": "MEDIO",
        "RIESGO ALTO": "ALTO",
        "ALTO": "ALTO",
        "INVIABLE SANITARIAMENTE": "INVIABLE SANITARIAMENTE",
        "INVIABLE": "INVIABLE SANITARIAMENTE",
        "SD": "SIN DATO",
        "S/D": "SIN DATO",
        "SIN DATO": "SIN DATO",
        "": "SIN DATO",
    }
    return mapping.get(text, "SIN DATO")


def parse_date(value):
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return pd.NaT
    return parsed.normalize()


def score_priority(row) -> float:
    risk = row["NIVEL_RIESGO_IRCA"]
    factor = RISK_FACTOR.get(risk, 0.5)
    students = float(row.get("ESTUDIANTES_ACTUAL", 0) or 0)
    concept = normalize_text(row.get("CONCEPTO_SANITARIO", ""))
    concept_factor = 1.2 if "DESFAVORABLE" in concept else (1.05 if "REQUERIMIENTO" in concept else 1.0)
    return round(factor * math.log10(students + 1) * concept_factor, 2)


def priority_band(score: float) -> str:
    if score >= 8:
        return "MUY ALTA"
    if score >= 6:
        return "ALTA"
    if score >= 4:
        return "MEDIA"
    return "BAJA"


def ensure_inputs() -> None:
    for folder in (RAW_DIR, CURRENT_DIR, DOCS_DIR, ARCHIVE_DIR):
        folder.mkdir(parents=True, exist_ok=True)

    missing = [path for path in (INPUT_IRCA, INPUT_SEDES, INPUT_MAESTRO_ACUEDUCTOS) if not path.exists()]
    if missing:
        print("[ERROR] No se encontraron los archivos fuente:")
        for path in missing:
            print(f"  - {path}")
        print("\nCopie los dos Excel en la carpeta raw y ejecute nuevamente.")
        raise SystemExit(1)


def backup_existing(path: Path) -> None:
    if path.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_folder = ARCHIVE_DIR / timestamp
        backup_folder.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, backup_folder / path.name)


# =============================================================================
# PROCESAMIENTO
# =============================================================================
def load_sources() -> tuple[pd.DataFrame, pd.DataFrame]:
    irca = pd.read_excel(INPUT_IRCA, dtype=object)
    sedes = pd.read_excel(INPUT_SEDES, dtype=object)

    irca.columns = [clean(column) for column in irca.columns]
    sedes.columns = [clean(column) for column in sedes.columns]

    irca = irca.loc[:, [c for c in irca.columns if c]]
    sedes = sedes.loc[:, [c for c in sedes.columns if c]]

    irca["ID_SEDE_IRCA"] = range(1, len(irca) + 1)
    sedes["ID_SEDESUESVALLE"] = range(1, len(sedes) + 1)

    for df in (irca, sedes):
        df["_COD"] = df["COD_DANE"].map(normalize_code)
        df["_MUN"] = df["MUNICIPIO"].map(normalize_text)
        df["_NOMBRE"] = df["NOMBRE_INS_PAE"].map(normalize_text)
        df["_KEY_NOMBRE"] = df["_MUN"] + "|" + df["_NOMBRE"]

    sedes["_LON"] = pd.to_numeric(sedes.get("LONGITUD"), errors="coerce")
    sedes["_LAT"] = pd.to_numeric(sedes.get("LATITUD"), errors="coerce")
    sedes["_COORD_OK"] = sedes["_LON"].notna() & sedes["_LAT"].notna()

    return irca, sedes


def choose_match(irca_row: pd.Series, sedes: pd.DataFrame):
    code = irca_row["_COD"]
    name_key = irca_row["_KEY_NOMBRE"]

    candidates_code = sedes[sedes["_COD"].eq(code)] if code else sedes.iloc[0:0]
    candidates_name = sedes[sedes["_KEY_NOMBRE"].eq(name_key)] if name_key != "|" else sedes.iloc[0:0]
    exact_code_name = candidates_code[candidates_code["_KEY_NOMBRE"].eq(name_key)]

    selected = None
    method = "SIN_COINCIDENCIA"
    status = "PENDIENTE"
    observation = ""

    if len(candidates_code) == 1:
        selected = candidates_code.iloc[0]
        method = "COD_DANE_UNICO"
        status = "CRUCE_OK"
    elif len(exact_code_name) == 1:
        selected = exact_code_name.iloc[0]
        method = "COD_DANE_DUPLICADO+NOMBRE_EXACTO"
        status = "CRUCE_OK_CONTROLADO"
    elif len(candidates_code) > 1:
        coords = candidates_code.loc[candidates_code["_COORD_OK"], ["_LAT", "_LON"]].drop_duplicates()
        if len(coords) == 1:
            selected = candidates_code[candidates_code["_COORD_OK"]].iloc[0]
            method = "COD_DANE_DUPLICADO_COORDENADA_COMUN"
            status = "VALIDAR_DUPLICADO"
            observation = "Código DANE duplicado en SedesUESVALLE; se asignó coordenada común."
        else:
            method = "COD_DANE_DUPLICADO_SIN_RESOLVER"
            status = "VALIDAR_DUPLICADO"
            observation = "Código DANE con múltiples candidatos; requiere selección manual."
    elif len(candidates_name) == 1:
        selected = candidates_name.iloc[0]
        method = "MUNICIPIO+NOMBRE_EXACTO"
        status = "CRUCE_RECUPERADO"
    elif len(candidates_name) > 1:
        coords = candidates_name.loc[candidates_name["_COORD_OK"], ["_LAT", "_LON"]].drop_duplicates()
        if len(coords) == 1:
            selected = candidates_name[candidates_name["_COORD_OK"]].iloc[0]
            method = "NOMBRE_DUPLICADO_COORDENADA_COMUN"
            status = "VALIDAR_DUPLICADO"
            observation = "Nombre duplicado; se asignó coordenada común."
        else:
            method = "NOMBRE_DUPLICADO_SIN_RESOLVER"
            status = "VALIDAR_DUPLICADO"
            observation = "Nombre exacto con múltiples candidatos; requiere selección manual."

    return selected, method, status, observation, len(candidates_code), len(candidates_name)


def consolidate(irca: pd.DataFrame, sedes: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    records = []
    controls = []

    for _, row in irca.iterrows():
        matched, method, status, observation, n_code, n_name = choose_match(row, sedes)
        has_match = matched is not None
        has_coordinates = bool(matched["_COORD_OK"]) if has_match else False

        out = {
            "ID_SEDE_IRCA": row["ID_SEDE_IRCA"],
            "COD_DANE": row["_COD"],
            "MUNICIPIO": clean(row.get("MUNICIPIO")),
            "NOMBRE_SEDE_IRCA": clean(row.get("NOMBRE_INS_PAE")),
            "ZONA_IRCA": normalize_text(row.get("ZONA")),
            "DIRECCION_IRCA": clean(row.get("DIRECCION_PAE")),
            "NOMBRE_ABREVIADO_IRCA": clean(row.get("NOMBREABREVIADO")),
            "TELEFONO_IRCA": clean(row.get("TELEFONO")),
            "ESTUDIANTES_ACTUAL": int(numeric(row.get("Número de Estudiantes_Actual"), 0)),
            "CONECTADA_ACUEDUCTO": normalize_text(row.get("Conectada_Acueducto (Si/No)")),
            "NOMBRE_ACUEDUCTO": clean(row.get("NOMBRE DE ACUEDUCTO")),
            "IRCA_ORIGINAL": clean(row.get("IRCA")),
            "NIVEL_RIESGO_IRCA": normalize_risk(row.get("IRCA")),
            "CODIGO_SISA_ACUEDUCTO": clean(row.get("codigo SISA de acueducto veredal")),
            "TRATAMIENTO_EN_ESCUELA": clean(row.get("TIPO DE TRATAMIENTO EN ESCUELA (SI EXISTE)\nFILTRO, PLANTA COMPACTA, ACUEDUCTO PROPIO")),
            "FUENTE_ABASTO": clean(row.get("FUENTE DE ABASTO(RIO, QUEBRADA, LLUVIA…)")),
            "CALIFICACION_BLOQUE_AGUA": clean(row.get("Calificacion Bloque Agua y Sanemiento (acta IVC)")),
            "CONCEPTO_SANITARIO": clean(row.get("CONCEPTO SANITARIO")),
            "FECHA_VISITA": parse_date(row.get("Fecha de visita")),
            "TIPO_CRUCE": method,
            "ESTADO_CRUCE": status,
            "ESTADO_GEORREFERENCIACION": "GEORREFERENCIADA" if has_coordinates else "SIN_COORDENADAS",
            "OBSERVACION_CRUCE": observation,
            "ID_SEDESUESVALLE": matched["ID_SEDESUESVALLE"] if has_match else "",
            "INSTITUCION_INS": clean(matched.get("INSTITUCION_INS")) if has_match else "",
            "SEDE_INS": clean(matched.get("SEDE_INS")) if has_match else "",
            "ESTADO_SEDE": clean(matched.get("ESTADO")) if has_match else "",
            "NATURALEZA": clean(matched.get("NATURALEZA")) if has_match else "",
            "JORNADA": clean(matched.get("JORNADA")) if has_match else "",
            "NIVEL_EDUCATIVO": clean(matched.get("NIVEL")) if has_match else "",
            "DIRECCION_MAPA": clean(matched.get("DIRECCION_MAPA")) if has_match else "",
            "LONGITUD": matched["_LON"] if has_match else None,
            "LATITUD": matched["_LAT"] if has_match else None,
            "BENEFICIARIOS_PAE_2025_1SI": int(numeric(matched.get("Beneficiarios_PAE_2025_1SI"), 0)) if has_match else 0,
            "TOTAL_GENERAL_INS": int(numeric(matched.get("Total_general_INS"), 0)) if has_match else 0,
        }
        out["SCORE_PRIORIDAD"] = score_priority(out)
        out["NIVEL_PRIORIDAD"] = priority_band(out["SCORE_PRIORIDAD"])
        records.append(out)

        controls.append({
            "ID_SEDE_IRCA": out["ID_SEDE_IRCA"],
            "COD_DANE": out["COD_DANE"],
            "MUNICIPIO": out["MUNICIPIO"],
            "NOMBRE_SEDE_IRCA": out["NOMBRE_SEDE_IRCA"],
            "CANDIDATOS_POR_CODIGO": n_code,
            "CANDIDATOS_POR_NOMBRE": n_name,
            "TIPO_CRUCE": method,
            "ESTADO_CRUCE": status,
            "ESTADO_GEORREFERENCIACION": out["ESTADO_GEORREFERENCIACION"],
            "OBSERVACION": observation,
        })

    return pd.DataFrame(records), pd.DataFrame(controls)



def create_aqueduct_map_layer() -> pd.DataFrame:
    """Prepara la capa de puntos de acueductos que será consumida por el HTML."""
    maestro = pd.read_csv(INPUT_MAESTRO_ACUEDUCTOS, encoding="utf-8-sig", sep=None, engine="python")
    maestro.columns = [clean(column) for column in maestro.columns]

    required = ["LATITUD", "LONGITUD", "CODIGO_SISTEMA", "MUNICIPIO", "PERSONA_PRESTADORA"]
    missing = [column for column in required if column not in maestro.columns]
    if missing:
        raise ValueError("El maestro de acueductos no contiene las columnas obligatorias: " + ", ".join(missing))

    for column in ["LATITUD", "LONGITUD", "POBLACION", "SUSCRIPTORES"]:
        if column in maestro.columns:
            maestro[column] = pd.to_numeric(maestro[column], errors="coerce")

    columns = [
        "ID_MAESTRO", "CODIGO_SISTEMA", "MUNICIPIO", "AMBITO", "TIPO_SISTEMA", "ARO",
        "PERSONA_PRESTADORA", "LOCALIDADES_ABASTECIDAS", "SUSCRIPTORES", "POBLACION",
        "TIPO_SISTEMA_TRATAMIENTO", "ALGUN_TIPO_TRATAMIENTO", "ESTADO_MAESTRO",
        "LATITUD", "LONGITUD", "VALIDACION_SISTEMA"
    ]
    available = [column for column in columns if column in maestro.columns]
    capa = maestro.loc[maestro["LATITUD"].notna() & maestro["LONGITUD"].notna(), available].copy()
    capa["MUNICIPIO"] = capa["MUNICIPIO"].map(normalize_text)
    return capa


def create_summaries(base: pd.DataFrame):
    risk = (
        base.groupby("NIVEL_RIESGO_IRCA", dropna=False)
        .agg(SEDES=("ID_SEDE_IRCA", "count"), ESTUDIANTES=("ESTUDIANTES_ACTUAL", "sum"))
        .reindex(RISK_ORDER, fill_value=0)
        .reset_index()
    )

    base_calc = base.copy()
    base_calc["SEDE_ALTO_INVIABLE"] = base_calc["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]).astype(int)
    base_calc["EST_MEDIO_MAS"] = base_calc["ESTUDIANTES_ACTUAL"].where(
        base_calc["NIVEL_RIESGO_IRCA"].isin(["MEDIO", "ALTO", "INVIABLE SANITARIAMENTE"]), 0
    )
    base_calc["EST_ALTO_INVIABLE"] = base_calc["ESTUDIANTES_ACTUAL"].where(
        base_calc["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]), 0
    )
    base_calc["GEORREFERENCIADA"] = base_calc["ESTADO_GEORREFERENCIACION"].eq("GEORREFERENCIADA").astype(int)

    counts = pd.crosstab(base_calc["MUNICIPIO"], base_calc["NIVEL_RIESGO_IRCA"])
    counts = counts.reindex(columns=RISK_ORDER, fill_value=0).reset_index()
    totals = (
        base_calc.groupby("MUNICIPIO", as_index=False)
        .agg(
            SEDES_TOTAL=("ID_SEDE_IRCA", "count"),
            ESTUDIANTES_TOTAL=("ESTUDIANTES_ACTUAL", "sum"),
            SEDES_ALTO_INVIABLE=("SEDE_ALTO_INVIABLE", "sum"),
            ESTUDIANTES_MEDIO_MAS=("EST_MEDIO_MAS", "sum"),
            ESTUDIANTES_ALTO_INVIABLE=("EST_ALTO_INVIABLE", "sum"),
            SEDES_GEORREFERENCIADAS=("GEORREFERENCIADA", "sum"),
        )
    )
    municipality = totals.merge(counts, on="MUNICIPIO", how="left")
    municipality = municipality.sort_values(["SEDES_ALTO_INVIABLE", "ESTUDIANTES_ALTO_INVIABLE"], ascending=False)

    aqueduct = (
        base_calc.assign(
            CODIGO_SISA_ACUEDUCTO=base_calc["CODIGO_SISA_ACUEDUCTO"].replace("", "SIN CODIGO SISA"),
            NOMBRE_ACUEDUCTO=base_calc["NOMBRE_ACUEDUCTO"].replace("", "SIN NOMBRE DE ACUEDUCTO"),
        )
        .groupby(["CODIGO_SISA_ACUEDUCTO", "NOMBRE_ACUEDUCTO"], dropna=False, as_index=False)
        .agg(
            MUNICIPIOS=("MUNICIPIO", lambda s: ", ".join(sorted(set(x for x in s if x)))),
            SEDES_TOTAL=("ID_SEDE_IRCA", "count"),
            ESTUDIANTES_TOTAL=("ESTUDIANTES_ACTUAL", "sum"),
            SEDES_ALTO_INVIABLE=("SEDE_ALTO_INVIABLE", "sum"),
            ESTUDIANTES_ALTO_INVIABLE=("EST_ALTO_INVIABLE", "sum"),
        )
        .sort_values(["SEDES_ALTO_INVIABLE", "ESTUDIANTES_ALTO_INVIABLE"], ascending=False)
    )
    return risk, municipality, aqueduct


# =============================================================================
# EXCEL DE CONTROL
# =============================================================================
def build_excel(base, control, risk, municipality, aqueduct, without_geo, validate_dates) -> None:
    for path in (OUT_XLSX,):
        backup_existing(path)

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        # Resumen inicial; se complementa con fórmulas y formato después.
        resumen_cruce = (
            base.groupby("ESTADO_CRUCE", as_index=False)
            .agg(REGISTROS=("ID_SEDE_IRCA", "count"))
            .sort_values("REGISTROS", ascending=False)
        )
        risk.to_excel(writer, sheet_name="RESUMEN_EJECUTIVO", index=False, startrow=14, startcol=0)
        resumen_cruce.to_excel(writer, sheet_name="RESUMEN_EJECUTIVO", index=False, startrow=14, startcol=4)
        municipality.head(10).to_excel(writer, sheet_name="RESUMEN_EJECUTIVO", index=False, startrow=24, startcol=0)

        base.to_excel(writer, sheet_name="BASE_CONSOLIDADA", index=False, startrow=2)
        control.to_excel(writer, sheet_name="CONTROL_CRUCE", index=False, startrow=2)
        base[base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"])].sort_values(
            ["SCORE_PRIORIDAD", "ESTUDIANTES_ACTUAL"], ascending=False
        ).to_excel(writer, sheet_name="SEDES_CRITICAS", index=False, startrow=2)
        without_geo.to_excel(writer, sheet_name="SEDES_SIN_COORDENADAS", index=False, startrow=2)
        municipality.to_excel(writer, sheet_name="RESUMEN_MUNICIPIO", index=False, startrow=2)
        aqueduct.to_excel(writer, sheet_name="RESUMEN_ACUEDUCTO", index=False, startrow=2)
        validate_dates.to_excel(writer, sheet_name="VALIDAR_FECHAS", index=False, startrow=2)

        rules = pd.DataFrame([
            ["Fuente", "Escuelas_IRCAS.xlsx", "Universo sanitario; cada registro se conserva.", "Base principal"],
            ["Fuente", "SedesUESVALLE.xlsx", "Coordenadas y atributos educativos complementarios.", "Cruce espacial"],
            ["Cruce", "COD_DANE_UNICO", "Coincidencia única por código DANE.", "Automático"],
            ["Cruce", "COD_DANE_DUPLICADO+NOMBRE_EXACTO", "Duplicado resuelto por nombre normalizado.", "Controlado"],
            ["Cruce", "MUNICIPIO+NOMBRE_EXACTO", "Recuperado cuando no se resuelve por código.", "Recuperado"],
            ["IRCA", "Normalización", "RIESGO ALTO/ALTO → ALTO; SD → SIN DATO.", "Simbología"],
            ["Prioridad", "SCORE_PRIORIDAD", "Factor de riesgo × log10(estudiantes + 1) × factor sanitario.", "Ranking"],
            ["Fecha", "Fecha de corte", f"Validar fechas posteriores a {FECHA_CORTE.isoformat()}.", "Control"],
        ], columns=["SECCION", "CAMPO_REGLA", "DESCRIPCION", "USO"])
        rules.to_excel(writer, sheet_name="DICCIONARIO_REGLAS", index=False, startrow=2)

    wb = load_workbook(OUT_XLSX)
    blue = "2E64D2"
    light_blue = "EAF1FF"
    dark = "14213D"
    border = Side(style="thin", color="D9E3F0")
    title_fill = PatternFill("solid", fgColor=blue)
    header_fill = PatternFill("solid", fgColor=blue)

    titles = {
        "RESUMEN_EJECUTIVO": "Tablero Sedes Educativas – Base consolidada de calidad del agua",
        "BASE_CONSOLIDADA": "Base consolidada sanitaria y espacial – una fila por sede evaluada",
        "CONTROL_CRUCE": "Control de cruce entre Escuelas_IRCAS y SedesUESVALLE",
        "SEDES_CRITICAS": "Sedes clasificadas en riesgo alto o inviable sanitariamente",
        "SEDES_SIN_COORDENADAS": "Sedes pendientes de georreferenciación para el mapa",
        "RESUMEN_MUNICIPIO": "Resumen municipal para KPIs y gráficos",
        "RESUMEN_ACUEDUCTO": "Acueductos abastecedores y sedes asociadas",
        "VALIDAR_FECHAS": "Fechas vacías o posteriores al corte",
        "DICCIONARIO_REGLAS": "Diccionario y reglas de consolidación",
    }

    for ws in wb.worksheets:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(ws.max_column, 4))
        cell = ws.cell(1, 1)
        cell.value = titles.get(ws.title, ws.title)
        cell.fill = title_fill
        cell.font = Font(color="FFFFFF", bold=True, size=15)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A4"

        header_row = 15 if ws.title == "RESUMEN_EJECUTIVO" else 3
        for cell in ws[header_row]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=border, right=border, top=border, bottom=border)
        ws.row_dimensions[header_row].height = 30
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

        for col in range(1, ws.max_column + 1):
            values = [clean(ws.cell(r, col).value) for r in range(1, min(ws.max_row, 60) + 1)]
            width = min(max(max((len(v) for v in values), default=10) + 2, 12), 42)
            ws.column_dimensions[get_column_letter(col)].width = width

    summary = wb["RESUMEN_EJECUTIVO"]
    summary["A3"] = "Objetivo: integrar la clasificación sanitaria asociada al agua que abastece las sedes educativas con su ubicación geográfica y caracterización educativa."
    summary.merge_cells("A3:L4")
    summary["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    summary["A3"].fill = PatternFill("solid", fgColor="F8FAFC")
    summary["A3"].font = Font(color="475467", size=10)

    kpis = [
        ("A6", "Total sedes evaluadas", len(base)),
        ("D6", "Sedes georreferenciadas", int((base["ESTADO_GEORREFERENCIACION"] == "GEORREFERENCIADA").sum())),
        ("G6", "Estudiantes reportados", int(base["ESTUDIANTES_ACTUAL"].sum())),
        ("J6", "Sedes alto/inviable", int(base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]).sum())),
        ("A10", "Estudiantes alto/inviable", int(base.loc[base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]), "ESTUDIANTES_ACTUAL"].sum())),
        ("D10", "Sedes sin IRCA", int((base["NIVEL_RIESGO_IRCA"] == "SIN DATO").sum())),
        ("G10", "Sedes sin coordenadas", int((base["ESTADO_GEORREFERENCIACION"] == "SIN_COORDENADAS").sum())),
        ("J10", "Cruces por validar/pendientes", int(base["ESTADO_CRUCE"].isin(["VALIDAR_DUPLICADO", "PENDIENTE"]).sum())),
    ]
    for start, label, value in kpis:
        col = summary[start].column
        row = summary[start].row
        summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
        summary.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 2)
        summary.cell(row, col).value = label
        summary.cell(row, col).fill = PatternFill("solid", fgColor=light_blue)
        summary.cell(row, col).font = Font(bold=True, color=dark, size=10)
        summary.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
        summary.cell(row + 1, col).value = value
        summary.cell(row + 1, col).font = Font(bold=True, color=dark, size=18)
        summary.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center")
        summary.cell(row + 1, col).number_format = "#,##0"

    chart = BarChart()
    chart.title = "Sedes educativas por nivel de riesgo"
    chart.y_axis.title = "Sedes"
    chart.x_axis.title = "Nivel de riesgo"
    data = Reference(summary, min_col=2, min_row=15, max_row=21)
    categories = Reference(summary, min_col=1, min_row=16, max_row=21)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 8
    chart.width = 16
    summary.add_chart(chart, "E32")

    # Tablas estructuradas en hojas de datos
    for ws in wb.worksheets:
        if ws.title == "RESUMEN_EJECUTIVO":
            continue
        ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
        table_name = re.sub(r"[^A-Za-z0-9]", "", ws.title.title()) + "Tabla"
        table = Table(displayName=table_name[:25], ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    wb.save(OUT_XLSX)


def main() -> None:
    print("=" * 88)
    print(" NORMALIZAR Y CONSOLIDAR SEDES EDUCATIVAS - IRCA | UESVALLE")
    print("=" * 88)
    print(f"Repositorio: {ROOT}")
    print(f"Fuentes    : {RAW_DIR}")
    print(f"Salidas    : {CURRENT_DIR}")
    print()

    ensure_inputs()
    irca, sedes = load_sources()
    base, control = consolidate(irca, sedes)
    risk, municipality, aqueduct = create_summaries(base)
    aqueduct_map = create_aqueduct_map_layer()

    without_geo = base[base["ESTADO_GEORREFERENCIACION"] == "SIN_COORDENADAS"].copy()
    validate_dates = base[
        base["FECHA_VISITA"].isna()
        | (base["FECHA_VISITA"].dt.date > FECHA_CORTE)
    ][["ID_SEDE_IRCA", "COD_DANE", "MUNICIPIO", "NOMBRE_SEDE_IRCA", "FECHA_VISITA"]].copy()
    validate_dates["VALIDACION"] = validate_dates["FECHA_VISITA"].apply(
        lambda x: "FECHA VACÍA O NO INTERPRETABLE" if pd.isna(x) else "FECHA POSTERIOR AL CORTE"
    )

    for output in (OUT_BASE_CSV, OUT_METADATA, OUT_RESUMEN_MUNICIPIO, OUT_RESUMEN_RIESGO, OUT_RESUMEN_ACUEDUCTO, OUT_CONTROL_CRUCE, OUT_SIN_GEO, OUT_VALIDAR_FECHAS, OUT_ACUEDUCTOS_MAPA):
        backup_existing(output)

    base.to_csv(OUT_BASE_CSV, index=False, encoding="utf-8-sig", sep=";")
    risk.to_csv(OUT_RESUMEN_RIESGO, index=False, encoding="utf-8-sig", sep=";")
    municipality.to_csv(OUT_RESUMEN_MUNICIPIO, index=False, encoding="utf-8-sig", sep=";")
    aqueduct.to_csv(OUT_RESUMEN_ACUEDUCTO, index=False, encoding="utf-8-sig", sep=";")
    control.to_csv(OUT_CONTROL_CRUCE, index=False, encoding="utf-8-sig", sep=";")
    without_geo.to_csv(OUT_SIN_GEO, index=False, encoding="utf-8-sig", sep=";")
    validate_dates.to_csv(OUT_VALIDAR_FECHAS, index=False, encoding="utf-8-sig", sep=";")
    aqueduct_map.to_csv(OUT_ACUEDUCTOS_MAPA, index=False, encoding="utf-8-sig", sep=";")

    metadata = {
        "fecha_generacion": datetime.now().isoformat(timespec="seconds"),
        "fecha_corte_validacion": FECHA_CORTE.isoformat(),
        "fuentes": [INPUT_IRCA.name, INPUT_SEDES.name, INPUT_MAESTRO_ACUEDUCTOS.name],
        "acueductos_georreferenciados": int(len(aqueduct_map)),
        "registros_sedes_irca": int(len(base)),
        "sedes_georreferenciadas": int((base["ESTADO_GEORREFERENCIACION"] == "GEORREFERENCIADA").sum()),
        "sedes_sin_coordenadas": int((base["ESTADO_GEORREFERENCIACION"] == "SIN_COORDENADAS").sum()),
        "sedes_alto_inviable": int(base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]).sum()),
        "estudiantes_alto_inviable": int(base.loc[base["NIVEL_RIESGO_IRCA"].isin(["ALTO", "INVIABLE SANITARIAMENTE"]), "ESTUDIANTES_ACTUAL"].sum()),
        "riesgo_sedes": {risk_value: int((base["NIVEL_RIESGO_IRCA"] == risk_value).sum()) for risk_value in RISK_ORDER},
        "riesgo_estudiantes": {risk_value: int(base.loc[base["NIVEL_RIESGO_IRCA"] == risk_value, "ESTUDIANTES_ACTUAL"].sum()) for risk_value in RISK_ORDER},
        "estado_cruce": {str(key): int(value) for key, value in Counter(base["ESTADO_CRUCE"]).items()},
    }
    with OUT_METADATA.open("w", encoding="utf-8") as file:
        json.dump(metadata, file, ensure_ascii=False, indent=2)

    build_excel(base, control, risk, municipality, aqueduct, without_geo, validate_dates)

    print("[OK] Proceso finalizado.")
    print(f"  Universo sedes                 : {len(base):,}".replace(",", "."))
    print(f"  Sedes georreferenciadas        : {(base['ESTADO_GEORREFERENCIACION'] == 'GEORREFERENCIADA').sum():,}".replace(",", "."))
    print(f"  Sedes sin coordenadas          : {(base['ESTADO_GEORREFERENCIACION'] == 'SIN_COORDENADAS').sum():,}".replace(",", "."))
    print(f"  Sedes alto o inviable          : {base['NIVEL_RIESGO_IRCA'].isin(['ALTO', 'INVIABLE SANITARIAMENTE']).sum():,}".replace(",", "."))
    print(f"  Estudiantes alto o inviable    : {base.loc[base['NIVEL_RIESGO_IRCA'].isin(['ALTO', 'INVIABLE SANITARIAMENTE']), 'ESTUDIANTES_ACTUAL'].sum():,}".replace(",", "."))
    print(f"  Cruces validar/pendientes      : {base['ESTADO_CRUCE'].isin(['VALIDAR_DUPLICADO', 'PENDIENTE']).sum():,}".replace(",", "."))
    print(f"  Acueductos georreferenciados   : {len(aqueduct_map):,}".replace(",", "."))
    print()
    print("Archivos generados:")
    print(f"  - {OUT_XLSX}")
    print(f"  - {OUT_BASE_CSV}")
    print(f"  - {OUT_METADATA}")
    print(f"  - {OUT_RESUMEN_MUNICIPIO}")
    print(f"  - {OUT_RESUMEN_RIESGO}")
    print(f"  - {OUT_RESUMEN_ACUEDUCTO}")
    print(f"  - {OUT_CONTROL_CRUCE}")
    print(f"  - {OUT_SIN_GEO}")
    print(f"  - {OUT_VALIDAR_FECHAS}")
    print(f"  - {OUT_ACUEDUCTOS_MAPA}")


if __name__ == "__main__":
    main()
